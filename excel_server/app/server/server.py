import io
import os
from contextlib import asynccontextmanager
from typing import List, Optional

import uvicorn
from app.file_extract_store.file_extract_store import FileExtractInfo, FileExtractStore
from app.file_store.file_store import FileStore, LocalFileStoreBackend, UserFile
from fastapi import Depends, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel

# --- Configuration ---
BASE_STORAGE_DIR = os.getenv("base_storage_dir", "./storage")

# --- Dependencies ---
file_store: Optional[FileStore] = None
file_extract_store: Optional[FileExtractStore] = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    global file_store, file_extract_store

    # Initialize FileStore
    fs_db_path = os.path.join(BASE_STORAGE_DIR, "file_store", "file_store_db.sqllite")
    fs_files_path = os.path.join(BASE_STORAGE_DIR, "file_store", "files")

    # Ensure directories exist
    os.makedirs(os.path.dirname(fs_db_path), exist_ok=True)

    file_store = FileStore(
        db_url=f"sqlite:///{fs_db_path}",
        backend=LocalFileStoreBackend(base_path=fs_files_path),
    )

    # Initialize FileExtractStore
    fes_db_path = os.path.join(
        BASE_STORAGE_DIR, "file_extract_store", "file_extract_store_db.sqllite"
    )

    # Ensure directories exist
    os.makedirs(os.path.dirname(fes_db_path), exist_ok=True)

    file_extract_store = FileExtractStore(db_url=f"sqlite:///{fes_db_path}")

    yield


app = FastAPI(lifespan=lifespan)

# --- CORS ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Adjust in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Models ---


class FileUpdate(BaseModel):
    original_filename: Optional[str] = None


class ExtractUpdate(BaseModel):
    payload: str


class SheetInfo(BaseModel):
    sheet_idx: int
    sheet_name: str
    extract: Optional[FileExtractInfo] = None


class FileDetailResponse(UserFile):
    sheets: List[SheetInfo]


# --- Helper Dependencies ---


def get_user_id():
    return "user_one"


def get_file_store() -> FileStore:
    if file_store is None:
        raise HTTPException(status_code=500, detail="FileStore not initialized")
    return file_store


def get_file_extract_store() -> FileExtractStore:
    if file_extract_store is None:
        raise HTTPException(status_code=500, detail="FileExtractStore not initialized")
    return file_extract_store


# --- Routes ---


@app.get("/")
def read_root():
    return {"message": "Hello, World!"}


@app.get("/files", response_model=List[UserFile])
def list_files(
    user_id: str = Depends(get_user_id), store: FileStore = Depends(get_file_store)
):
    return store.list_files(user_id)


@app.get("/files/{file_id}", response_model=FileDetailResponse)
def get_file_details(
    file_id: str,
    user_id: str = Depends(get_user_id),
    f_store: FileStore = Depends(get_file_store),
    fe_store: FileExtractStore = Depends(get_file_extract_store),
):
    try:
        # Get file metadata and content
        user_file, content = f_store.get_file(user_id, file_id)

        # Parse Excel to get sheets
        # Using io.BytesIO because openpyxl expects a file-like object
        try:
            workbook = load_workbook(filename=io.BytesIO(content), read_only=True)
            sheet_names = workbook.sheetnames
        except Exception as e:
            # If we can't parse it as excel, maybe return empty sheets or error?
            # Assuming strictly excel files for this usecase
            print(f"Error parsing excel file: {e}")
            sheet_names = []

        sheets = []
        for idx, name in enumerate(sheet_names):
            # Get latest extract for this sheet
            extract = fe_store.get_latest(user_id, file_id, idx)
            sheets.append(SheetInfo(sheet_idx=idx, sheet_name=name, extract=extract))

        return FileDetailResponse(**user_file.model_dump(), sheets=sheets)

    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/files/{file_id}", response_model=UserFile)
def delete_file(
    file_id: str,
    user_id: str = Depends(get_user_id),
    store: FileStore = Depends(get_file_store),
):
    try:
        return store.delete_file(user_id, file_id)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/analyze/{file_id}/{sheet_idx}")
def analyze_file_sheet(
    file_id: str,
    sheet_idx: int,
    user_id: str = Depends(get_user_id),
    f_store: FileStore = Depends(get_file_store),
    fe_store: FileExtractStore = Depends(get_file_extract_store),
):
    # Check if file exists
    try:
        f_store.get_file_metadata(user_id, file_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found")

    # TODO: Invoke AI workflow here.
    # For now, we'll just simulate a "processing" state or similar if needed.
    # The instructions don't specify what this should return or do exactly other than "Invokes the AI workflow".
    # We could trigger a background task.

    return {"message": "Analysis started", "file_id": file_id, "sheet_idx": sheet_idx}


@app.post("/update/{file_id}/{sheet_idx}", response_model=FileExtractInfo)
def update_extract(
    file_id: str,
    sheet_idx: int,
    update: ExtractUpdate,
    user_id: str = Depends(get_user_id),
    fe_store: FileExtractStore = Depends(get_file_extract_store),
):
    try:
        return fe_store.add_extract(user_id, file_id, sheet_idx, update.payload)
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/upload", response_model=UserFile)
async def upload_file(
    file: UploadFile = File(...),
    user_id: str = Depends(get_user_id),
    store: FileStore = Depends(get_file_store),
) -> UserFile:
    content = await file.read()
    filename = file.filename or "unknown"
    return store.create_file(user_id, filename, content)


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8080)
