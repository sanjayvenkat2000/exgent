import io

from app.domain import SheetData
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def convert_excel_to_sheet_data(excel_bytes: bytes) -> list[tuple[str, SheetData]]:
    workbook = load_workbook(
        filename=io.BytesIO(excel_bytes), read_only=True, data_only=True
    )
    sheets_data = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # In read_only mode, max_column might not be accurate until read,
        # so we read all rows first to determine dimensions and content.
        raw_rows = []
        max_col_count = 0

        for row in sheet.iter_rows(values_only=True):
            raw_rows.append(row)
            # row is a tuple of values
            if len(row) > max_col_count:
                max_col_count = len(row)

        # Prepare data for this sheet
        sheet_output = []

        # 1. Create header row: " ", "A", "B", ...
        # If the sheet is empty (max_col_count == 0), we might just return empty or specific handling.
        # Assuming typical sheet has at least 1 column if it has data.
        if max_col_count > 0:
            header_row = [" "] + [
                get_column_letter(i) for i in range(1, max_col_count + 2)
            ]
            sheet_output.append(header_row)

            # 2. Iterate rows, add row number, empty column, and convert values
            for idx, row_values in enumerate(raw_rows, start=1):
                # Start with row number and empty scratchpad column
                processed_row = [str(idx), ""]

                # Append values, handling None and padding if necessary
                for col_idx in range(max_col_count):
                    if col_idx < len(row_values):
                        val = row_values[col_idx]
                        processed_row.append(str(val) if val is not None else "")
                    else:
                        # Pad with empty string if row is shorter than max width
                        processed_row.append("")

                sheet_output.append(processed_row)

        sheets_data.append((sheet_name, SheetData(data=sheet_output)))

    return sheets_data
